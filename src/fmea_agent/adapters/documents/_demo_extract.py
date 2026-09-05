"""Bounded local text/BOM extraction behind D2's document adapter."""

import csv
import io
from dataclasses import dataclass
from typing import Literal
from zipfile import BadZipFile, ZipFile

ErrorCode = Literal[
    "UNSUPPORTED_FORMAT",
    "LIMIT_EXCEEDED",
    "NO_TEXT",
    "ENCRYPTED",
    "INVALID_BOM",
    "PARTIAL_MODEL",
    "INVALID_MODEL",
]
MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_TEXT_CHARS = 30_000
MAX_BOM_ROWS = 200
BOM_HEADER = ("item_id", "parent_id", "name", "quantity", "unit", "source_element_id")


class DemoInputError(ValueError):
    """Safe application-facing diagnostic; never include parser exception text."""

    def __init__(self, code: ErrorCode, message: str) -> None:
        self.code = code
        super().__init__(message)


@dataclass
class TextBudget:
    count: int = 0

    def add(self, text: str) -> None:
        self.count += len(text)
        if self.count > MAX_TEXT_CHARS:
            raise DemoInputError("LIMIT_EXCEEDED", "输入提取文本合计超过 30,000 字符。")


def decode_text(raw: bytes) -> str:
    if raw.lstrip().startswith((b"%PDF-", b"PK\x03\x04", b"\xd0\xcf\x11\xe0")):
        raise DemoInputError("UNSUPPORTED_FORMAT", "文本扩展名与文件内容不符。")
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise DemoInputError("UNSUPPORTED_FORMAT", "文本必须使用 UTF-8 编码。") from exc
    if any(ord(char) < 32 and char not in "\t\r\n" for char in text):
        raise DemoInputError("UNSUPPORTED_FORMAT", "文本包含不支持的二进制控制字符。")
    return text


def extract_design(raw: bytes, suffix: str, budget: TextBudget) -> list[tuple[str, str]]:
    if suffix != ".pdf":
        text = decode_text(raw)
        budget.add(text)
        if not text.strip():
            raise DemoInputError("NO_TEXT", "设计说明没有可提取文本。")
        return [
            (f"line={number}", line)
            for number, line in enumerate(text.splitlines(), 1)
            if line.strip()
        ]
    if not raw.startswith(b"%PDF-"):
        raise DemoInputError("UNSUPPORTED_FORMAT", "PDF 扩展名与内容不符。")
    # Optional dependency is loaded only on the PDF path.
    from pypdf import PdfReader

    try:
        reader = PdfReader(io.BytesIO(raw), strict=True)
        if reader.is_encrypted:
            raise DemoInputError("ENCRYPTED", "不支持加密 PDF。")
        if len(reader.pages) > 20:
            raise DemoInputError("LIMIT_EXCEEDED", "PDF 超过 20 页。")
        if not reader.pages:
            raise DemoInputError("NO_TEXT", "PDF 没有可提取文本的页面。")
        result = []
        for number, page in enumerate(reader.pages, 1):
            text = page.extract_text()
            budget.add(text)
            if not text.strip():
                raise DemoInputError("NO_TEXT", f"PDF 第 {number} 页没有文本；不支持空页或扫描页。")
            result.append((f"page={number}", text))
        return result
    except DemoInputError:
        raise
    except Exception as exc:
        raise DemoInputError("UNSUPPORTED_FORMAT", "PDF 内容无法解析。") from exc


def _xlsx_rows(raw: bytes) -> list[tuple[int, list[str]]]:
    if not raw.startswith(b"PK\x03\x04"):
        raise DemoInputError("UNSUPPORTED_FORMAT", "XLSX 扩展名与内容不符。")
    try:
        with ZipFile(io.BytesIO(raw)) as archive:
            # Bound the archive before openpyxl inflates shared strings/styles.
            if sum(info.file_size for info in archive.infolist()) > 25 * 1024 * 1024:
                raise DemoInputError("LIMIT_EXCEEDED", "XLSX 展开内容超过 25 MiB。")
            content_types = archive.read("[Content_Types].xml")
            if (
                b"macroEnabled" in content_types
                or b"spreadsheetml.sheet.main+xml" not in content_types
            ):
                raise DemoInputError("UNSUPPORTED_FORMAT", "只支持无宏的 XLSX 工作簿。")
    except (BadZipFile, KeyError) as exc:
        raise DemoInputError("UNSUPPORTED_FORMAT", "文件不是有效的 XLSX 工作簿。") from exc
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    book = None
    try:
        book = load_workbook(io.BytesIO(raw), read_only=True, data_only=False, keep_links=False)
        if "BOM" not in book.sheetnames:
            raise DemoInputError("INVALID_BOM", "XLSX 必须包含 BOM 工作表。")
        sheet = book["BOM"]
        sheet.reset_dimensions()  # Do not trust declared dimensions that hide data/formulas.
        rows = []
        for number, cells in enumerate(sheet.iter_rows(), 1):
            if any(cell.data_type == "f" for cell in cells):
                raise DemoInputError("INVALID_BOM", f"BOM 第 {number} 行含公式。")
            values = ["" if cell.value is None else str(cell.value) for cell in cells]
            while len(values) > len(BOM_HEADER) and not values[-1]:
                values.pop()
            if not any(values):
                continue
            rows.append((number, values))
            if len(rows) > MAX_BOM_ROWS + 1:
                raise DemoInputError("LIMIT_EXCEEDED", "BOM 超过 200 数据行。")
        return rows
    except DemoInputError:
        raise
    except Exception as exc:
        raise DemoInputError("UNSUPPORTED_FORMAT", "XLSX 内容无法解析。") from exc
    finally:
        if book is not None:
            book.close()


def extract_bom(
    raw: bytes,
    suffix: str,
    budget: TextBudget,
) -> list[tuple[str, dict[str, str]]]:
    if suffix == ".csv":
        text = decode_text(raw)
        budget.add(text)
        try:
            reader = csv.reader(io.StringIO(text, newline=""), strict=True)
            rows = []
            for number, row in enumerate(reader, 1):
                if any(row):
                    rows.append((number, row))
                if len(rows) > MAX_BOM_ROWS + 1:
                    raise DemoInputError("LIMIT_EXCEEDED", "BOM 超过 200 数据行。")
        except csv.Error as exc:
            raise DemoInputError("INVALID_BOM", "CSV 内容无法解析。") from exc
    else:
        rows = _xlsx_rows(raw)
        for _, row in rows:
            budget.add("\n".join(row))
    if not rows or tuple(rows[0][1]) != BOM_HEADER:
        raise DemoInputError("INVALID_BOM", "BOM 必须包含规定的六列，且顺序一致。")
    result = []
    seen = set()
    for number, values in rows[1:]:
        if len(values) != len(BOM_HEADER):
            raise DemoInputError("INVALID_BOM", f"BOM 第 {number} 行列数不符。")
        if any(value.lstrip().startswith(("=", "+", "-", "@")) for value in values):
            raise DemoInputError("INVALID_BOM", f"BOM 第 {number} 行含公式前缀。")
        record = dict(zip(BOM_HEADER, values, strict=True))
        if not all(record[key].strip() for key in ["item_id", "name", "unit", "source_element_id"]):
            raise DemoInputError("INVALID_BOM", f"BOM 第 {number} 行缺少必填值。")
        if record["item_id"] in seen:
            raise DemoInputError("INVALID_BOM", "BOM item_id 重复。")
        seen.add(record["item_id"])
        locator = f"sheet=BOM&row={number}" if suffix == ".xlsx" else f"row={number}"
        result.append((locator, record))
    return result
